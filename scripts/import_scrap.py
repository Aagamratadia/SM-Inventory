"""
Usage:
  MONGODB_URI="mongodb+srv://user:pass@cluster/dbname?opts" python scripts/import_scrap.py [path-to-xlsx]

Notes:
- Reads the first worksheet of the Excel file.
- Maps headers case-insensitively: Name, Category, Quantity, Price, Notes, Vendor Name, Vendor Contact, Vendor Email, Vendor Address, Item ID, Shape, Scrapped On.
- Upserts by `name`. Sets `isScrap = True` and `scrappedAt`.
- Requires env MONGODB_URI. If your URI lacks a database name, set MONGODB_DB too.
Dependencies:
  pip install openpyxl pymongo
"""
from __future__ import annotations

import os
import sys
import math
from datetime import datetime
from typing import Any, Dict, List, Optional

from pymongo import MongoClient
from pymongo.collection import Collection

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl. Install with: pip install openpyxl")
    sys.exit(1)


def load_env_from_file(file_path: str) -> None:
    """Load KEY=VALUE lines from a dotenv file into os.environ without overwriting existing vars."""
    if not os.path.exists(file_path):
        return
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                # handle quoted values and ignore export prefix
                if line.startswith("export "):
                    line = line[len("export "):]
                if "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        # Non-fatal: continue without raising
        pass

def load_env_local() -> None:
    """Load .env.local (and fallback .env) from project root (parent of scripts/)."""
    scripts_dir = os.path.dirname(__file__)
    project_root = os.path.abspath(os.path.join(scripts_dir, os.pardir))
    env_local = os.path.join(project_root, ".env.local")
    env_file = os.path.join(project_root, ".env")
    load_env_from_file(env_local)
    load_env_from_file(env_file)


def to_number(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return None
        return float(val)
    try:
        f = float(str(val).strip())
        if math.isnan(f) or math.isinf(f):
            return None
        return f
    except Exception:
        return None


essential_header_keys = {
    "name": ["name", "item", "item name", "product", "product name", "itemname"],
    "category": ["category"],
    "quantity": ["quantity", "qty"],
    "price": ["price", "unit price"],
    "notes": ["notes", "note", "remark", "remarks"],
    "vendorname": ["vendorname", "vendor name", "vendor"],
    "vendorContact": ["vendorcontact", "vendor contact", "contact"],
    "vendorEmail": ["vendoremail", "vendor email", "email"],
    "vendorAddress": ["vendoraddress", "vendor address", "address"],
    "itemId": ["itemid", "item id", "sku"],
    "shape": ["shape"],
    "scrappedAt": ["scrapped on", "scrapped at", "date", "scrap date"],
}


def normalize_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def parse_date(v: Any) -> Optional[datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    # openpyxl may return numbers for Excel dates depending on format. Attempt parse.
    # Try several common string formats.
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"): 
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        # Fallback: ISO parser
        return datetime.fromisoformat(s)
    except Exception:
        return None


def read_rows(xlsx_path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    if not wb.sheetnames:
        raise RuntimeError("No sheets found in the workbook")
    ws = wb[wb.sheetnames[0]]

    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([c for c in row])
    if not rows:
        return []

    # Auto-detect header row within the first 10 rows
    expected_tokens = set(t for keys in essential_header_keys.values() for t in keys)
    header_idx = None
    for i in range(min(10, len(rows))):
        cells = rows[i]
        norm = [str(h).strip().lower() if h is not None else "" for h in cells]
        if any(col in expected_tokens for col in norm):
            header_idx = i
            header_norm = norm
            break
    if header_idx is None:
        # No recognizable header; use single-column fallback (treat first column as 'name')
        print("No recognizable header found. Using single-column fallback: first column = name")
        items: List[Dict[str, Any]] = []
        for r in rows:
            if not r:
                continue
            first = r[0]
            if first is None or str(first).strip() == "":
                continue
            items.append({"name": str(first).strip()})
        return items

    # Debug: show detected header
    print(f"Detected header row at index {header_idx}: {header_norm}")

    def get_cell(row_vals: List[Any], keys: List[str]) -> Any:
        keys_norm = [k.lower() for k in keys]
        for idx, col in enumerate(header_norm):
            if col in keys_norm:
                return row_vals[idx] if idx < len(row_vals) else None
        return None

    items: List[Dict[str, Any]] = []
    for r in rows[header_idx + 1:]:
        # skip entirely empty rows
        if all(v in (None, "") for v in r):
            continue
        obj: Dict[str, Any] = {}
        for out_key, keys in essential_header_keys.items():
            obj[out_key] = get_cell(r, keys)
        items.append(obj)
    return items


def connect_collection() -> Collection:
    mongo_uri = os.getenv("MONGODB_URI")
    if not mongo_uri:
        print("Error: MONGODB_URI is not set in the environment (load from .env.local or set it).")
        sys.exit(1)

    client = MongoClient(mongo_uri)
    db = client.get_default_database()

    if db is None:
        db_name = os.getenv("MONGODB_DB")
        if not db_name:
            print("Error: Database name not found in URI and MONGODB_DB not set.")
            print("Provide a db name in the MONGODB_URI (e.g., .../mydb) or set MONGODB_DB.")
            sys.exit(1)
        db = client[db_name]

    # Mongoose model name 'Item' defaults to 'items' collection
    return db["items"]


def upsert_items(col: Collection, items: List[Dict[str, Any]]) -> Dict[str, int]:
    inserted = 0
    updated = 0
    failed = 0

    for idx, raw in enumerate(items, start=1):
        name = normalize_str(raw.get("name"))
        category = normalize_str(raw.get("category")) or name  # fallback
        # Default vendorname to item name to avoid unique index conflicts on vendorname
        vendorname = normalize_str(raw.get("vendorname")) or name
        vendor_contact = normalize_str(raw.get("vendorContact"))
        vendor_email = normalize_str(raw.get("vendorEmail"))
        vendor_address = normalize_str(raw.get("vendorAddress"))
        item_id = normalize_str(raw.get("itemId"))
        shape = normalize_str(raw.get("shape"))
        notes = normalize_str(raw.get("notes"))

        qty_num = to_number(raw.get("quantity"))
        quantity = int(qty_num) if qty_num is not None else 0
        price_val = to_number(raw.get("price"))
        scrapped_at = parse_date(raw.get("scrappedAt")) or datetime.now()

        if not name:
            failed += 1
            print(f"Row {idx}: Skipped (missing name)")
            continue

        doc = {
            "category": category,
            "name": name,
            "vendorname": vendorname,
            "vendorContact": vendor_contact,
            "vendorEmail": vendor_email,
            "vendorAddress": vendor_address,
            "itemId": item_id,
            "shape": shape,
            "price": price_val,
            "quantity": quantity,
            "totalQuantity": quantity,
            "notes": notes,
            "isScrap": True,
            "scrappedAt": scrapped_at,
        }

        try:
            res = col.update_one({"name": name}, {"$set": doc}, upsert=True)
            if res.upserted_id is not None:
                inserted += 1
            elif res.modified_count > 0:
                updated += 1
            else:
                # matched but nothing changed
                updated += 0
        except Exception as e:
            failed += 1
            print(f"Row {idx}: Failed -> {e}")

    return {"inserted": inserted, "updated": updated, "failed": failed}


def main() -> None:
    # Load env from project root before doing anything else
    load_env_local()

    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), "scrap.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"Error: File not found at {xlsx_path}")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    rows = read_rows(xlsx_path)
    print(f"Rows found: {len(rows)}")

    col = connect_collection()
    print("Connected to MongoDB")

    stats = upsert_items(col, rows)
    print(f"\nDone. Inserted: {stats['inserted']}, Updated: {stats['updated']}, Failed: {stats['failed']}")


if __name__ == "__main__":
    main()
